from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Asset, Company, InventoryCampaign
from app.models.asset import AssetStatus
from app.auth import require_user
from app.models.user import User
from app.templates_ctx import templates
from app.routers.assets_router import EQUIPMENT_KIND_LABELS

router = APIRouter(prefix="", tags=["reports"])

# Типы техники для отчёта «Светофор» (компы, без мониторов/принтеров/телефонов)
TRAFFIC_LIGHT_KINDS = ("desktop", "nettop", "laptop", "server")

STATUS_LABELS = {
    "active": "Активно", "inactive": "Неактивно",
    "maintenance": "На обслуживании", "retired": "Списано",
}


def _age_years(manufacture_date: date | None) -> float | None:
    if not manufacture_date:
        return None
    today = date.today()
    delta = (today - manufacture_date).days
    return round(delta / 365.25, 1)


def _traffic_color(age_years: float | None, threshold_years: int) -> str:
    if age_years is None:
        return "secondary"  # серый — нет даты
    if age_years < 3:
        return "success"   # зелёный
    if age_years < threshold_years:
        return "warning"  # жёлтый
    return "danger"       # красный


# Порядок сортировки по цвету: красный → жёлтый → зелёный → серый
COLOR_SORT_ORDER = {"danger": 0, "warning": 1, "success": 2, "secondary": 3}
# Заливка для Excel (светофор)
EXCEL_FILLS = {
    "danger": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),   # светлый красный
    "warning": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # светлый жёлтый
    "success": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),  # светлый зелёный
    "secondary": PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
}


@router.get("/reports", name="reports")
async def reports(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_user),
):
    total_assets = (await db.execute(select(func.count(Asset.id)))).scalar() or 0
    by_status = await db.execute(
        select(Asset.status, func.count(Asset.id)).group_by(Asset.status)
    )
    status_counts = dict(by_status.all())
    total_campaigns = (await db.execute(select(func.count(InventoryCampaign.id)))).scalar() or 0
    return templates.TemplateResponse(
        "reports.html",
        {
            "request": request,
            "user": current_user,
            "total_assets": total_assets,
            "status_counts": status_counts,
            "status_enum": AssetStatus,
            "status_labels": STATUS_LABELS,
            "total_campaigns": total_campaigns,
        },
    )


@router.get("/reports/traffic-light", name="reports_traffic_light")
async def reports_traffic_light(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_user),
    company_id: str | None = Query(None, description="Организация"),
    threshold_years: int = Query(5, ge=1, le=20, description="Порог устаревания (лет), красный цвет"),
):
    company_id_int = None
    if company_id and str(company_id).strip():
        try:
            company_id_int = int(company_id.strip())
        except ValueError:
            pass
    companies_result = await db.execute(select(Company).order_by(Company.name))
    companies = list(companies_result.scalars().all())
    q = (
        select(Asset)
        .where(Asset.equipment_kind.in_(TRAFFIC_LIGHT_KINDS))
        .options(selectinload(Asset.company))
        .order_by(Asset.company_id, Asset.name)
    )
    if company_id_int is not None:
        q = q.where(Asset.company_id == company_id_int)
    result = await db.execute(q)
    assets = list(result.scalars().all())
    rows = []
    for a in assets:
        age = _age_years(getattr(a, "manufacture_date", None))
        color = _traffic_color(age, threshold_years)
        rows.append({"asset": a, "age_years": age, "color": color})
    rows.sort(key=lambda r: (COLOR_SORT_ORDER.get(r["color"], 99), (r["asset"].name or "").lower()))
    return templates.TemplateResponse(
        "reports_traffic_light.html",
        {
            "request": request,
            "user": current_user,
            "companies": companies,
            "company_id": company_id_int,
            "threshold_years": threshold_years,
            "rows": rows,
            "equipment_kind_labels": EQUIPMENT_KIND_LABELS,
        },
    )


def _status_label_for_color(color: str, threshold_years: int) -> str:
    if color == "success":
        return "до 3 лет"
    if color == "warning":
        return f"3–{threshold_years} лет"
    if color == "danger":
        return f"старше {threshold_years}"
    return "нет даты"


@router.get("/reports/traffic-light/export.xlsx", name="reports_traffic_light_export")
async def reports_traffic_light_export(
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_user),
    company_id: str | None = Query(None, description="Организация"),
    threshold_years: int = Query(5, ge=1, le=20, description="Порог устаревания (лет)"),
):
    company_id_int = None
    if company_id and str(company_id).strip():
        try:
            company_id_int = int(company_id.strip())
        except ValueError:
            pass
    q = (
        select(Asset)
        .where(Asset.equipment_kind.in_(TRAFFIC_LIGHT_KINDS))
        .options(selectinload(Asset.company))
        .order_by(Asset.company_id, Asset.name)
    )
    if company_id_int is not None:
        q = q.where(Asset.company_id == company_id_int)
    result = await db.execute(q)
    assets = list(result.scalars().all())
    rows = []
    for a in assets:
        age = _age_years(getattr(a, "manufacture_date", None))
        color = _traffic_color(age, threshold_years)
        rows.append({"asset": a, "age_years": age, "color": color})
    rows.sort(key=lambda r: (COLOR_SORT_ORDER.get(r["color"], 99), (r["asset"].name or "").lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Светофор"
    headers = ["Название", "Тип", "Организация", "Дата выпуска", "Возраст (лет)", "Статус"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, r in enumerate(rows, 2):
        a = r["asset"]
        ws.cell(row=row_idx, column=1, value=a.name or "")
        ws.cell(row=row_idx, column=2, value=EQUIPMENT_KIND_LABELS.get(a.equipment_kind, a.equipment_kind or "—"))
        ws.cell(row=row_idx, column=3, value=a.company.name if a.company else "—")
        ws.cell(row=row_idx, column=4, value=a.manufacture_date.strftime("%d.%m.%Y") if getattr(a, "manufacture_date", None) else "—")
        ws.cell(row=row_idx, column=5, value=r["age_years"] if r["age_years"] is not None else "—")
        ws.cell(row=row_idx, column=6, value=_status_label_for_color(r["color"], threshold_years))
        fill = EXCEL_FILLS.get(r["color"], EXCEL_FILLS["secondary"])
        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).fill = fill
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=traffic_light.xlsx"},
    )
