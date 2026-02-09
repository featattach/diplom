from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Asset, InventoryCampaign, InventoryItem


# Кириллические заголовки для экспорта оборудования
ASSET_EXPORT_HEADERS = [
    "ID", "Название", "Модель", "Тип техники", "Организация", "Серийный номер",
    "Категория", "Расположение", "Статус", "Последняя активность", "Дата создания",
]

def export_assets_xlsx(assets: list[Asset]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборудование"

    for col, h in enumerate(ASSET_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    status_labels = {"active": "Активно", "inactive": "Неактивно", "maintenance": "На обслуживании", "retired": "Списано"}
    for row, asset in enumerate(assets, 2):
        company_name = ""
        if getattr(asset, "company", None) and asset.company:
            company_name = asset.company.name
        ws.cell(row=row, column=1, value=asset.id)
        ws.cell(row=row, column=2, value=asset.name)
        ws.cell(row=row, column=3, value=getattr(asset, "model", None) or "")
        ws.cell(row=row, column=4, value=getattr(asset, "equipment_kind", None) or "")
        ws.cell(row=row, column=5, value=company_name)
        ws.cell(row=row, column=6, value=asset.serial_number or "")
        ws.cell(row=row, column=7, value=asset.asset_type or "")
        ws.cell(row=row, column=8, value=asset.location or "")
        ws.cell(row=row, column=9, value=status_labels.get(asset.status.value, asset.status.value) if asset.status else "")
        ws.cell(row=row, column=10, value=asset.last_seen_at.isoformat() if asset.last_seen_at else "")
        ws.cell(row=row, column=11, value=asset.created_at.isoformat() if asset.created_at else "")

    for col in range(1, len(ASSET_EXPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_inventory_campaign_xlsx(campaign: InventoryCampaign, items: list[InventoryItem]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    ws.cell(row=1, column=1, value="Campaign")
    ws.cell(row=1, column=2, value=campaign.name)
    ws.cell(row=2, column=1, value="Started")
    ws.cell(row=2, column=2, value=campaign.started_at.isoformat() if campaign.started_at else "")
    ws.cell(row=3, column=1, value="Finished")
    ws.cell(row=3, column=2, value=campaign.finished_at.isoformat() if campaign.finished_at else "")

    headers = ["ID", "Asset ID", "Asset name", "Expected location", "Found", "Found at", "Notes"]
    start_row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for row, item in enumerate(items, start_row + 1):
        asset = item.asset
        ws.cell(row=row, column=1, value=item.id)
        ws.cell(row=row, column=2, value=item.asset_id or "")
        ws.cell(row=row, column=3, value=asset.name if asset else "")
        ws.cell(row=row, column=4, value=item.expected_location or "")
        ws.cell(row=row, column=5, value="Yes" if item.found else "No")
        ws.cell(row=row, column=6, value=item.found_at.isoformat() if item.found_at else "")
        ws.cell(row=row, column=7, value=item.notes or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
