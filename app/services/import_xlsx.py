"""
Импорт оборудования из Excel. Первая строка — заголовки.
Поддерживаются: название, модель, тип техники, серийный номер, расположение, статус,
категория, описание, организация, пользователь; плюс технические поля: CPU, ОЗУ, тип/объём диска,
IP адрес, мат. плата, ОС, блок питания, диагональ экрана/монитора, разрешение, юниты (U), дата выпуска.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.models.asset import AssetStatus

# Заголовки шаблона (и варианты для распознавания в загружаемом файле)
IMPORT_HEADERS = [
    "Название",
    "Модель",
    "Тип техники",
    "Серийный номер",
    "Расположение",
    "Статус",
    "Категория",
    "Описание",
    "Организация",
    "Пользователь (кто использует)",
    "CPU",
    "ОЗУ",
    "Тип диска",
    "Объём диска",
    "IP адрес",
    "Мат. плата",
    "ОС",
    "Блок питания",
    "Диагональ экрана",
    "Разрешение экрана",
    "Диагональ монитора",
    "Юниты (U)",
    "Дата выпуска",
]

# Варианты названий столбцов (старые инвентарки могут называть иначе)
HEADER_ALIASES = {
    "название": "Название",
    "имя": "Название",
    "наименование": "Название",
    "модель": "Модель",
    "тип техники": "Тип техники",
    "тип": "Тип техники",
    "вид техники": "Тип техники",
    "серийный номер": "Серийный номер",
    "серийный": "Серийный номер",
    "s/n": "Серийный номер",
    "расположение": "Расположение",
    "локация": "Расположение",
    "кабинет": "Расположение",
    "место": "Расположение",
    "статус": "Статус",
    "категория": "Категория",
    "описание": "Описание",
    "организация": "Организация",
    "компания": "Организация",
    "пользователь (кто использует)": "Пользователь (кто использует)",
    "пользователь": "Пользователь (кто использует)",
    "ответственный": "Пользователь (кто использует)",
    "фio": "Пользователь (кто использует)",
    "cpu": "CPU",
    "процессор": "CPU",
    "озу": "ОЗУ",
    "ram": "ОЗУ",
    "память": "ОЗУ",
    "тип диска": "Тип диска",
    "объём диска": "Объём диска",
    "диск": "Тип диска",
    "ip адрес": "IP адрес",
    "ip": "IP адрес",
    "сетевая карта": "IP адрес",
    "мат. плата": "Мат. плата",
    "материнская плата": "Мат. плата",
    "ос": "ОС",
    "блок питания": "Блок питания",
    "диагональ экрана": "Диагональ экрана",
    "разрешение экрана": "Разрешение экрана",
    "разрешение": "Разрешение экрана",
    "диагональ монитора": "Диагональ монитора",
    "юниты (u)": "Юниты (U)",
    "юниты": "Юниты (U)",
    "u": "Юниты (U)",
    "дата выпуска": "Дата выпуска",
    "дата производства": "Дата выпуска",
}

# Значения статуса в файле -> AssetStatus
STATUS_MAP = {
    "активно": AssetStatus.active,
    "active": AssetStatus.active,
    "неактивно": AssetStatus.inactive,
    "inactive": AssetStatus.inactive,
    "на обслуживании": AssetStatus.maintenance,
    "maintenance": AssetStatus.maintenance,
    "списано": AssetStatus.retired,
    "retired": AssetStatus.retired,
}

# Тип техники: подпись в файле -> value в БД (equipment_kind)
EQUIPMENT_KIND_LABELS_TO_VALUE = {
    "системный блок": "desktop",
    "десктоп": "desktop",
    "desktop": "desktop",
    "неттоп": "nettop",
    "nettop": "nettop",
    "ноутбук": "laptop",
    "laptop": "laptop",
    "монитор": "monitor",
    "monitor": "monitor",
    "мфу": "mfu",
    "mfu": "mfu",
    "принтер": "printer",
    "printer": "printer",
    "сканер": "scanner",
    "scanner": "scanner",
    "коммутатор": "switch",
    "switch": "switch",
    "сервер": "server",
    "server": "server",
    "sip-телефон": "sip_phone",
    "sip телефон": "sip_phone",
    "sip_phone": "sip_phone",
    "моноблок": "desktop",  # маппим в desktop при отсутствии monoblock в модели
}


def _normalize_header(cell_value: Any) -> str | None:
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    return s if s else None


def _normalize_cell(cell_value: Any) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _map_header(raw: str) -> str | None:
    if not raw:
        return None
    lower = raw.lower().strip()
    return HEADER_ALIASES.get(lower) or (raw if raw in IMPORT_HEADERS else None)


def _parse_date(s: str) -> date | None:
    """Парсит дату из строки: YYYY-MM-DD или DD.MM.YYYY."""
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(s: str) -> int | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def build_import_template_xlsx() -> BytesIO:
    """Генерирует пустой шаблон Excel с одной строкой заголовков для импорта."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборудование"
    for col, h in enumerate(IMPORT_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_import_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Парсит загруженный Excel. Первая строка — заголовки.
    Возвращает (список строк как dict с ключами полей Asset, список ошибок).
    """
    errors = []
    rows_out = []
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            errors.append("В файле нет листа.")
            return [], errors
    except Exception as e:
        errors.append(f"Не удалось открыть файл: {e}")
        return [], errors

    # Собираем заголовки первой строки и маппим в стандартные имена
    header_row = []
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        val = ws.cell(row=1, column=c).value
        raw = _normalize_header(val)
        mapped = _map_header(raw) if raw else None
        header_row.append(mapped)

    if not any(header_row):
        errors.append("Не найдена строка заголовков (первая строка).")
        return [], errors

    if "Название" not in header_row:
        errors.append("Обязательный столбец «Название» не найден. Переименуйте столбец с названием оборудования в «Название».")

    # Читаем данные
    for row_idx in range(2, (ws.max_row or 1) + 1):
        row_data = {}
        for col_idx, std_name in enumerate(header_row):
            if not std_name:
                continue
            cell_val = ws.cell(row=row_idx, column=col_idx + 1).value
            row_data[std_name] = _normalize_cell(cell_val)

        name = row_data.get("Название", "").strip()
        if not name:
            continue  # пустая строка — пропускаем

        # Преобразуем в поля модели
        status_val = row_data.get("Статус", "").strip().lower()
        status = STATUS_MAP.get(status_val) if status_val else AssetStatus.active

        kind_raw = row_data.get("Тип техники", "").strip().lower()
        equipment_kind = None
        for label, value in EQUIPMENT_KIND_LABELS_TO_VALUE.items():
            if label == kind_raw or kind_raw == value:
                equipment_kind = value
                break
        # Неизвестный тип не подставляем — в карточке будет «—», можно поправить вручную

        rows_out.append({
            "name": name,
            "model": row_data.get("Модель", "") or None,
            "equipment_kind": equipment_kind or None,
            "serial_number": row_data.get("Серийный номер", "") or None,
            "location": row_data.get("Расположение", "") or None,
            "status": status,
            "asset_type": row_data.get("Категория", "") or None,
            "description": row_data.get("Описание", "") or None,
            "company_name": row_data.get("Организация", "").strip() or None,
            "current_user": row_data.get("Пользователь (кто использует)", "") or None,
            "cpu": row_data.get("CPU", "") or None,
            "ram": row_data.get("ОЗУ", "") or None,
            "disk1_type": row_data.get("Тип диска", "") or None,
            "disk1_capacity": row_data.get("Объём диска", "") or None,
            "network_card": row_data.get("IP адрес", "") or None,
            "motherboard": row_data.get("Мат. плата", "") or None,
            "os": row_data.get("ОС", "") or None,
            "power_supply": row_data.get("Блок питания", "") or None,
            "screen_diagonal": row_data.get("Диагональ экрана", "") or None,
            "screen_resolution": row_data.get("Разрешение экрана", "") or None,
            "monitor_diagonal": row_data.get("Диагональ монитора", "") or None,
            "rack_units": _parse_int(row_data.get("Юниты (U)", "")),
            "manufacture_date": _parse_date(row_data.get("Дата выпуска", "")),
        })

    return rows_out, errors
