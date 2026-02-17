"""Формирование отчётов: светофор, экспорт оборудования в Excel с метаданными."""
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.constants import equipment_kind_label
from app.models import Asset
from app.repositories import asset_repo
from app.schemas.reports import EquipmentReportFilter, TrafficLightReportFilter
from app.services.export_xlsx import export_assets_xlsx
from app.utils.asset_helpers import is_asset_inactive

COLOR_SORT_ORDER = {"danger": 0, "warning": 1, "success": 2, "secondary": 3}
EXCEL_FILLS = {
    "danger": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "warning": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    "success": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "secondary": PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
}


def _age_years(manufacture_date: date | None) -> float | None:
    if not manufacture_date:
        return None
    today = date.today()
    delta = (today - manufacture_date).days
    return round(delta / 365.25, 1)


def _traffic_color(age_years: float | None, threshold_years: int) -> str:
    if age_years is None:
        return "secondary"
    if age_years < 3:
        return "success"
    if age_years < threshold_years:
        return "warning"
    return "danger"


def _status_label_for_color(color: str, threshold_years: int) -> str:
    if color == "success":
        return "до 3 лет"
    if color == "warning":
        return f"3–{threshold_years} лет"
    if color == "danger":
        return f"старше {threshold_years}"
    return "нет даты"


def build_traffic_light_rows(assets: list[Asset], threshold_years: int) -> list[dict[str, Any]]:
    """Строит список строк для отчёта «Светофор» (asset, age_years, color), отсортированный по цвету и имени."""
    rows = []
    for a in assets:
        age = _age_years(getattr(a, "manufacture_date", None))
        color = _traffic_color(age, threshold_years)
        rows.append({"asset": a, "age_years": age, "color": color})
    rows.sort(key=lambda r: (COLOR_SORT_ORDER.get(r["color"], 99), (r["asset"].name or "").lower()))
    return rows


def build_traffic_light_xlsx(rows: list[dict], threshold_years: int) -> BytesIO:
    """Формирует Excel-файл отчёта «Светофор» с заливкой по цветам."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Светофор"
    headers = ["Название", "Тип", "Организация", "Дата выпуска", "Возраст (лет)", "Статус"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, r in enumerate(rows, 2):
        a = r["asset"]
        ws.cell(row=row_idx, column=1, value=a.name or "")
        ws.cell(row=row_idx, column=2, value=equipment_kind_label(a.equipment_kind))
        ws.cell(row=row_idx, column=3, value=a.company.name if a.company else "—")
        ws.cell(row=row_idx, column=4, value=a.manufacture_date.strftime("%d.%m.%Y") if getattr(a, "manufacture_date", None) else "—")
        ws.cell(row=row_idx, column=5, value=r["age_years"] if r["age_years"] is not None else "—")
        ws.cell(row=row_idx, column=6, value=_status_label_for_color(r["color"], threshold_years))
        fill = EXCEL_FILLS.get(r["color"], EXCEL_FILLS["secondary"])
        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).fill = fill
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 18
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _add_metadata_sheet(
    workbook_buf: BytesIO,
    generated_at: datetime,
    generated_by: str,
    filters_description: str,
) -> BytesIO:
    """Добавляет лист «Метаданные» в начало книги: кто сформировал, когда, какие фильтры."""
    wb = load_workbook(workbook_buf)
    meta = wb.create_sheet("Метаданные", 0)
    meta.cell(row=1, column=1, value="Сформировано")
    meta.cell(row=1, column=2, value=generated_at.strftime("%Y-%m-%d %H:%M:%S UTC"))
    meta.cell(row=2, column=1, value="Пользователь")
    meta.cell(row=2, column=2, value=generated_by or "—")
    meta.cell(row=3, column=1, value="Параметры отчёта")
    meta.cell(row=3, column=2, value=filters_description or "—")
    for r in range(1, 4):
        meta.cell(row=r, column=1).font = Font(bold=True)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _equipment_filter_description(f: EquipmentReportFilter) -> str:
    parts = []
    if f.name:
        parts.append(f"название: {f.name}")
    if f.status:
        parts.append(f"статус: {f.status}")
    if f.inactive_by_activity:
        parts.append("неактивные по активности")
    if f.equipment_kind:
        parts.append(f"тип: {f.equipment_kind}")
    if f.location:
        parts.append(f"расположение: {f.location}")
    if f.company_id:
        parts.append(f"организация ID: {f.company_id}")
    parts.append(f"сортировка: {f.sort_value()}")
    return "; ".join(parts) if parts else "без фильтров"


async def export_equipment_xlsx(
    db: AsyncSession,
    filters: EquipmentReportFilter,
    generated_by: str,
    generated_at: datetime,
) -> BytesIO:
    """Экспорт отчёта «Оборудование» в Excel с метаданными (кто, когда, фильтры)."""
    sort_val = filters.sort_value()
    assets = await asset_repo.get_assets_list(
        db,
        name=filters.name,
        status=filters.status,
        inactive_by_activity=filters.inactive_by_activity,
        equipment_kind=filters.equipment_kind,
        location=filters.location,
        company_id=filters.company_id,
        sort=sort_val,
    )
    if filters.inactive_by_activity:
        assets = [a for a in assets if is_asset_inactive(a)]
    buf = export_assets_xlsx(assets)
    return _add_metadata_sheet(
        buf,
        generated_at,
        generated_by,
        _equipment_filter_description(filters),
    )


async def export_traffic_light_xlsx(
    db: AsyncSession,
    filters: TrafficLightReportFilter,
    generated_by: str,
    generated_at: datetime,
) -> BytesIO:
    """Экспорт отчёта «Светофор» в Excel с метаданными."""
    assets = await asset_repo.get_traffic_light_assets(db, filters.company_id)
    rows = build_traffic_light_rows(assets, filters.threshold_years)
    buf = build_traffic_light_xlsx(rows, filters.threshold_years)
    desc = f"Организация ID: {filters.company_id or 'все'}; порог устаревания: {filters.threshold_years} лет"
    return _add_metadata_sheet(buf, generated_at, generated_by, desc)
